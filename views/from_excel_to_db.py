from openpyxl import load_workbook

from settings import DbSetting
from models.models import CarcassMarketPrice, create_table


class DbInsert(object):
    """"""

    def insert_carcass(self):
        """ carcassテーブルにsummaryをinsert """
        wb_summary = load_workbook('豚枝肉相場_Summary.xlsx')
        ws_summary = wb_summary['Sheet1']

        # DB接続
        db_setting = DbSetting()
        session = db_setting.get_db_session()

        engine = create_table()

        # Summaryからデータを取得する
        for row in range(3, ws_summary.max_row + 1):
            market_date = ws_summary.cell(row, 1).value
            if market_date is None or market_date == 'None':
                break

            # 全農値
            nationwide_slaughter = ws_summary.cell(row, 2).value
            zennoh_high_price = ws_summary.cell(row, 3).value
            zennoh_middle_price = ws_summary.cell(row, 4).value
            # Tokyo
            tokyo_high_price = ws_summary.cell(row, 5).value
            tokyo_middle_price = ws_summary.cell(row, 6).value
            tokyo_ordinary_price = ws_summary.cell(row, 7).value
            tokyo_outside_price = ws_summary.cell(row, 8).value
            tokyo_head_count = ws_summary.cell(row, 9).value
            # Saitama
            saitama_high_price = ws_summary.cell(row, 10).value
            saitama_middle_price = ws_summary.cell(row, 11).value
            saitama_ordinary_price = ws_summary.cell(row, 12).value
            saitama_outside_price = ws_summary.cell(row, 13).value
            saitama_head_count = ws_summary.cell(row, 14).value
            # Yokohama
            yokohama_high_price = ws_summary.cell(row, 15).value
            yokohama_middle_price = ws_summary.cell(row, 16).value
            yokohama_ordinary_price = ws_summary.cell(row, 17).value
            yokohama_outside_price = ws_summary.cell(row, 18).value
            yokohama_head_count = ws_summary.cell(row, 19).value
            # Osaka
            osaka_high_price = ws_summary.cell(row, 20).value
            osaka_middle_price = ws_summary.cell(row, 21).value
            osaka_ordinary_price = ws_summary.cell(row, 22).value
            osaka_outside_price = ws_summary.cell(row, 23).value
            osaka_head_count = ws_summary.cell(row, 24).value

            model = CarcassMarketPrice(
                market_date=market_date,
                nationwide_slaughter=nationwide_slaughter,
                zennoh_high_price=zennoh_high_price,
                zennoh_middle_price=zennoh_middle_price,
                tokyo_high_price=tokyo_high_price,
                tokyo_middle_price=tokyo_middle_price,
                tokyo_ordinary_price=tokyo_ordinary_price,
                tokyo_outside_price=tokyo_outside_price,
                tokyo_head_count=tokyo_head_count,
                saitama_high_price=saitama_high_price,
                saitama_middle_price=saitama_middle_price,
                saitama_ordinary_price=saitama_ordinary_price,
                saitama_outside_price=saitama_outside_price,
                saitama_head_count=saitama_head_count,
                yokohama_high_price=yokohama_high_price,
                yokohama_middle_price=yokohama_middle_price,
                yokohama_ordinary_price=yokohama_ordinary_price,
                yokohama_outside_price=yokohama_outside_price,
                yokohama_head_count=yokohama_head_count,
                osaka_high_price=osaka_high_price,
                osaka_middle_price=osaka_middle_price,
                osaka_ordinary_price=osaka_ordinary_price,
                osaka_outside_price=osaka_outside_price,
                osaka_head_count=osaka_head_count
            )

            session.add(model)
            session.commit()

        print('summaryのDBインサートが完了しました。')

        # 処理終了後にsummaryの値がある行を削除する
        for row in ws_summary.iter_rows(min_row=3):
            for cell in row:
                cell.value = None

        wb_summary.save('豚枝肉相場_Summary.xlsx')
        wb_summary.close()

        # DBと切断
        session.close()

        # エンジン破棄
        db_setting.dispose_db_engine(engine)
