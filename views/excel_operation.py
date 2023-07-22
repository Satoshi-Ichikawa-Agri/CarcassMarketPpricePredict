"""Excel Operation module"""
import re
from openpyxl import load_workbook

from const import Const


class ExcelOperation(object):
    """Base Excel Operation"""

    def __init__(self):
        """コンストラクタ"""
        pass

    def __del__(self):
        """デストラクタ"""
        print("ExcelOperationオブジェクトを破棄します。")

    # ===== Common Parts =========================================
    def read_excel(self, workbook, worksheet):
        """read excel book and sheet

        Args:
            workbook (str): 対象のExcelBookのPath
            worksheet (str): 対象のExcelSheet名
        """
        workbook = load_workbook(workbook)
        worksheet = workbook[worksheet]

        return workbook, worksheet

    def get_cell_value(self, worksheet, row, column) -> str:
        """get cell value

        Args:
            worksheet (): 対象のExcelBookのシートオブジェクト
            row (int): 取得する値のセルの行番号
            column (int): 取得する値のセルの列番号
        """
        if row < 0 or column < 0:
            return ""

        value = str(worksheet.cell(row=row, column=column).value)

        if Const.is_null_or_empty(value):
            return ""

        return value

    def set_cell_value(self, worksheet, row, column, value):
        """set cell value

        Args:
            worksheet (): 対象のExcelBookのシートオブジェクト
            row (int): 入力したいセルの行番号
            column (int): 入力したいセルの列番号
            value (): 入力する値
        """
        worksheet.cell(row=row, column=column).value = value

    def save_excel(self, workbook, file_path=None):
        """Save ExcelBook and Close"""
        workbook.save(file_path)

    def save_and_close_excel(self, workbook, file=None, save_flag=False):
        """Save ExcelBook and Close"""
        if save_flag:
            workbook.save(file)

        workbook.close()

    def excel_to_pdf(self):
        """Excel to PDF"""
        pass

    # ===== Data Cleansing =========================================
    def get_market_date(self, worksheet, row) -> str:
        """取得判断フラグとして、A列の市場日付を取得する """
        return self.get_cell_value(worksheet, row, 1)

    def type_conversion(self, value) -> int | str:
        """type conversion"""
        if not Const.is_null_or_empty(value):
            value = int(value)

        return value

    def set_value_type_converted(self, worksheet, row, column, value):
        """ 対象セルに値をセットする """
        # 値を型変換する
        value_converted = self.type_conversion(value)
        # 値をセルにセット
        self.set_cell_value(worksheet, row, column, value_converted)

    def set_value_of_date(self, worksheet, row, column, value):
        """ 対象セルに値をセットする(date型の値限定) """
        self.set_cell_value(worksheet, row, column, value)

    def remove_date(self, value):
        """ 豚肉相場一覧表_yyyymm.xlsxの「全国と畜頭数」列の値が
        数値と日付の情報となっているので、日付を削除する処理
        変更前: 68800 (2023/02/28)
        変更後: 68800
        """
        # 正規表現
        regex = r"\(\d{4}/\d{1,2}/\d{1,2}\)"
        # 正規表現で示して値だけ削除し、さらに空白を削除する
        removed_value = str.strip(re.sub(regex, "", value))

        return removed_value
