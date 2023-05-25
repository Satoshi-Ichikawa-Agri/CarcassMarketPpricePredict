import os
import shutil
import re
from openpyxl import load_workbook

from models.models import CarcassMarketPriceExcel
from const import Const


class DataCleansing(object):
    """ Data Cleansing """
    
    def __init__(self, file_date):
        self.file_date = file_date
        
        # 対象のExcelファイルを指定する
        self.wb_original = load_workbook(f'download/豚肉相場一覧表_{ self.file_date }.xlsx')
        self.wb_summary = load_workbook('豚枝肉相場_Summary.xlsx')
        # 対象のシートを指定する
        self.ws_original = self.wb_original[f'豚肉相場一覧表_{ self.file_date }']
        self.ws_summary = self.wb_summary['Sheet1']


    def get_market_date(self, row):
        """取得判断フラグとして、A列の市場日付を取得する """
        return self.get_value(row, 1)


    def get_value(self, row, column):
        """ 対象セルの値を取得する """
        if column < 0 or row < 0:
            return ''
        value = str(self.ws_original.cell(row=row, column=column).value)
        
        if Const.is_null_or_empty(value):
            return ''
        
        return value


    def set_value(self, row, column, value):
        """ 対象セルに値をセットする """
        self.ws_summary.cell(row=row, column=column).value = self.type_conversion(value)


    def set_value_of_date(self, row, column, value):
        """ 対象セルに値をセットする(date型の値限定) """
        self.ws_summary.cell(row=row, column=column).value = value


    def remove_date(self, value):
        """ 豚肉相場一覧表_yyyymm.xlsxの「全国と畜頭数」列の値が
        数値と日付の情報となっているので、日付を削除する処理
        変更前: 68800 (2023/02/28)
        変更後: 68800
        """
        # 正規表現
        regex = r'\(\d{4}/\d{1,2}/\d{1,2}\)'
        # 正規表現で示して値だけ削除し、さらに空白を削除する
        removed_value = str.strip(re.sub(regex, '', value))
        
        return removed_value
    
    
    def type_conversion(self, value):
        """ 型変換関数
        valueがない → str
        valueがある → int
        """
        if not Const.is_null_or_empty(value):
            value = int(value)

        return value
    
    
    def summary_copy(self):
        """ Summaryファイルをoutputフォルダにコピーする """
        summary_file_path = os.path.join(Const.WORKSPACE_DIR, '豚枝肉相場_Summary.xlsx')
        output_file_path = os.path.join(Const.OUTPUT_DIR, f'豚枝肉相場_Summary_{ self.file_date }.xlsx')
        shutil.copy2(summary_file_path, output_file_path)


    def data_cleansing_process(self):
        """ データクレンジング処理 """

        model_list = []
        
        for row in range(1, 30):
            market_date = self.get_market_date(row)
            
            if '豚肉相場' in market_date:
                continue
            if Const.is_null_or_empty(market_date):
                continue
            if '全農建値' in market_date:
                break
            
            market_date = str(self.file_date) + Const.date_replace(market_date)
            model = CarcassMarketPriceExcel()
            model.market_date = market_date
            
            model_list.append(model)
            
        print('データ件数: {}件'.format(len(model_list)))
        
        # 元データの値を取得する
        for i, model in enumerate(model_list):
            model: CarcassMarketPriceExcel = model
            model.index = i
            row = i + 6
            
            # 全農建値
            model.zennoh_high_price = self.get_value(row, 3)
            model.zennoh_middle_price = self.get_value(row, 5)
            model.nationwide_slaughter = self.remove_date(self.get_value(row, 8))
            # Tokyo
            model.tokyo_high_price = self.get_value(row, 11)
            model.tokyo_middle_price = self.get_value(row, 13)
            model.tokyo_ordinary_price = self.get_value(row, 15)
            model.tokyo_outside_price = self.get_value(row, 17)
            model.tokyo_head_count = self.get_value(row, 19)
            # Saitama
            model.saitama_high_price = self.get_value(row, 22)
            model.saitama_middle_price = self.get_value(row, 24)
            model.saitama_ordinary_price = self.get_value(row, 26)
            model.saitama_outside_price = self.get_value(row, 28)
            model.saitama_head_count = self.get_value(row, 30)
            # Yokohama
            model.yokohama_high_price = self.get_value(row, 33)
            model.yokohama_middle_price = self.get_value(row, 35)
            model.yokohama_ordinary_price = self.get_value(row, 37)
            model.yokohama_outside_price = self.get_value(row, 39)
            model.yokohama_head_count = self.get_value(row, 41)
            # Osaka
            model.osaka_high_price = self.get_value(row, 44)
            model.osaka_middle_price = self.get_value(row, 46)
            model.osaka_ordinary_price = self.get_value(row, 48)
            model.osaka_outside_price = self.get_value(row, 50)
            model.osaka_head_count = self.get_value(row, 52)
        
        # Summaryにセットする
        for model in model_list:
            model: CarcassMarketPriceExcel = model
            row = model.index + 3
            
            self.set_value_of_date(row, 1, Const.from_str_to_date(model.market_date))
            # 全農建値
            self.set_value(row, 2, model.nationwide_slaughter)
            self.set_value(row, 3, model.zennoh_high_price)
            self.set_value(row, 4, model.zennoh_middle_price)
            # Tokyo
            self.set_value(row, 5, model.tokyo_high_price)
            self.set_value(row, 6, model.tokyo_middle_price)
            self.set_value(row, 7, model.tokyo_ordinary_price)
            self.set_value(row, 8, model.tokyo_outside_price)
            self.set_value(row, 9, model.tokyo_head_count)
            # Saitama
            self.set_value(row, 10, model.saitama_high_price)
            self.set_value(row, 11, model.saitama_middle_price)
            self.set_value(row, 12, model.saitama_ordinary_price)
            self.set_value(row, 13, model.saitama_outside_price)
            self.set_value(row, 14, model.saitama_head_count)
            # Yokohama
            self.set_value(row, 15, model.yokohama_high_price)
            self.set_value(row, 16, model.yokohama_middle_price)
            self.set_value(row, 17, model.yokohama_ordinary_price)
            self.set_value(row, 18, model.yokohama_outside_price)
            self.set_value(row, 19, model.yokohama_head_count)
            # Osaka
            self.set_value(row, 20, model.osaka_high_price)
            self.set_value(row, 21, model.osaka_middle_price)
            self.set_value(row, 22, model.osaka_ordinary_price)
            self.set_value(row, 23, model.osaka_outside_price)
            self.set_value(row, 24, model.osaka_head_count)
        
        self.wb_summary.save('豚枝肉相場_Summary.xlsx')
        
        self.wb_original.close()
        self.wb_summary.close()
        
        self.summary_copy()
